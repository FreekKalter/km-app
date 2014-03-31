from openpyxl import Workbook,load_workbook
from openpyxl.cell import get_column_letter

#wb = Workbook()

#dest_filename = r'empty_book.xlsx'

#ws = wb.active

#ws.title = "range names"

#for col_idx in xrange(1, 40):
    #col = get_column_letter(col_idx)
    #for row in xrange(1, 600):
        #ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

#ws = wb.create_sheet()

#ws.title = 'Pi'

#ws['F5'] = 3.14

#wb.save(filename = dest_filename)

wb = load_workbook(filename = r'uren_template.xlsx')
sheet_ranges = wb['range names']

sheet_ranges['B7'].value = 10

wb.save(filename = 'new.xlsx')
#print sheet_ranges['A6'].value # D18

